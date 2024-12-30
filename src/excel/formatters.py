# src/excel/formatters.py

from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Example color fills
LABEL_FILL = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")  # Light blue
DATA_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")   # Cornsilk

# Fonts
TITLE_FONT = Font(name="Times New Roman", size=14, bold=True, italic=True)
LABEL_FONT = Font(name="Times New Roman", size=10, bold=True, italic=True)
DATA_TNR_FONT = Font(name="Times New Roman", size=10)
DATA_TNR_ITALIC_FONT = Font(name="Times New Roman", size=10, italic=True)
DATA_TNR_BOLD_FONT = Font(name="Times New Roman", bold=True)
DATA_ARIAL_FONT = Font(name="Arial", size=10)
DATA_ARIAL_BOLD_FONT = Font(name="Arial", size=10, bold=True)
DATA_ARIAL_ITALIC_FONT = Font(name="Arial", size=10, italic=True)

# Borders
THIN_BORDER_SIDE = Side(style='thin', color='000000')
THIN_BORDER = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE
)
THICK_BORDER_SIDE = Side(style='thick', color='000000')
THICK_BORDER = Border(
    left=THICK_BORDER_SIDE, right=THICK_BORDER_SIDE, top=THICK_BORDER_SIDE, bottom=THICK_BORDER_SIDE
)

def title_fill_range(ws, row_number, left_col, right_col):
    for cc in range(left_col, right_col + 1):
        cell = ws.cell(row_number, cc)
        if cell.fill.patternType is None:
            cell.fill = LABEL_FILL

def apply_table_border(ws, row, start_col, end_col):
    """
    Applies a thin border around a group of cells in one row, from start_col to end_col.
    """
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        existing_border = cell.border
        new_border = Border(
            top=THIN_BORDER_SIDE,
            bottom=THIN_BORDER_SIDE,
            left=existing_border.left or THIN_BORDER_SIDE if col == start_col else existing_border.left,
            right=existing_border.right or THIN_BORDER_SIDE if col == end_col else existing_border.right,
            diagonal=existing_border.diagonal,
            diagonal_direction=existing_border.diagonal_direction,
            outline=existing_border.outline,
            vertical=existing_border.vertical,
            horizontal=existing_border.horizontal,
        )
        cell.border = new_border
