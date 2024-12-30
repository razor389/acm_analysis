# src/excel/generator.py

import os
import sys
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import textwrap
import re

from .formatters import (
    LABEL_FILL,
    DATA_FILL,
    TITLE_FONT,
    LABEL_FONT,
    DATA_TNR_FONT,
    DATA_TNR_ITALIC_FONT,
    DATA_TNR_BOLD_FONT,
    DATA_ARIAL_FONT,
    DATA_ARIAL_ITALIC_FONT,
    THIN_BORDER,
    THICK_BORDER,
    title_fill_range,
    apply_table_border
)

def create_or_append_xls(xls_filename):
    file_exists = os.path.exists(xls_filename)
    if file_exists:
        # Append mode with overlay if sheet exists
        writer = pd.ExcelWriter(xls_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    else:
        writer = pd.ExcelWriter(xls_filename, engine='openpyxl', mode='w')
    return writer, file_exists

def format_workbook(writer):
    """Remove gridlines from all worksheets."""
    for sheetname in writer.book.sheetnames:
        ws = writer.book[sheetname]
        ws.sheet_view.showGridLines = False

def load_final_output(ticker):
    file_path = os.path.join("output", f"{ticker}_yoy_consolidated.json")
    if not os.path.exists(file_path):
        print(f"Error: {file_path} does not exist.")
        sys.exit(1)
    with open(file_path, "r") as f:
        data = json.load(f)
    return data

def generate_excel_for_ticker_year(ticker: str, year: int):
    """
    Reproduce the old main usage:
    Generate the Excel file for the given ticker/year => ./output/{ticker}.{YY}.2.xlsx
    """
    ticker = ticker.upper()
    year_2_digits = str(year)[-2:]
    xls_filename = os.path.join("output", f"{ticker}.{year_2_digits}.2.xlsx")

    # 1) Load JSON data
    final_output = load_final_output(ticker)

    # 2) Create or append XLS
    writer, file_exists = create_or_append_xls(xls_filename)

    # 3) Write data
    _write_all_sheets(writer, final_output)

    # 4) Formatting
    format_workbook(writer)

    # 5) Save
    writer.close()
    print(f"Data for {ticker} written to {xls_filename} successfully.")

def _write_all_sheets(writer, final_output):
    """Write each sheet from final_output just like old code did."""
    _write_summary_sheet(writer, final_output)
    _write_company_description(writer, final_output)
    _write_analyses_sheet(writer, final_output)
    _write_profit_desc_sheet(writer, final_output)
    _write_balance_sheet_sheet(writer, final_output)
    _write_studies_sheet(writer, final_output)
    _write_qualities_sheet(writer, final_output)

#
# The next set of private _write_* methods are basically your old
# write_summary_sheet, write_company_description, etc. from gen_excel.py.
# The only difference is they're now prefixed with '_' to indicate internal usage,
# and we import style objects from formatters.py. 
# For brevity, we show a couple of them:
#

def _write_summary_sheet(writer, final_output):
    wb = writer.book
    if 'Summary' not in wb.sheetnames:
        wb.create_sheet('Summary')
    ws = wb['Summary']

    summary_data = final_output["summary"]
    company_name = summary_data["company_name"]
    exchange = summary_data["exchange"]
    symbol = summary_data["symbol"]
    description = summary_data["description"]

    # Example usage
    combined_title = f"{company_name.upper()} ({exchange}) - {symbol}"
    ws['E1'] = combined_title
    ws['E1'].font = TITLE_FONT

    ws['A4'] = "Company Description"
    ws['A4'].font = LABEL_FONT

    # Wrap lines
    wrapped_lines = textwrap.wrap(description, width=150)
    start_row = 5
    col = 2  # column B
    for i, line in enumerate(wrapped_lines):
        cell = ws.cell(row=start_row + i, column=col, value=line)
        # Additional styling if needed

def _write_company_description(writer, final_output):
    # ...
    pass

def _write_analyses_sheet(writer, final_output):
    # ...
    pass

def _write_profit_desc_sheet(writer, final_output):
    # ...
    pass

def _write_balance_sheet_sheet(writer, final_output):
    # ...
    pass

def _write_studies_sheet(writer, final_output):
    # ...
    pass

def _write_qualities_sheet(writer, final_output):
    # ...
    pass
